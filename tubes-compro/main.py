import calendar
import io
import json
import os
import random
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

random.seed(42)
np.random.seed(42)



CATEGORY = "MEDIUM"

_today = date.today()
MONTH = _today.month
YEAR  = _today.year
DAYS  = calendar.monthrange(YEAR, MONTH)[1]

SHIFT_DEF = {
    "S1": {"label": "Pagi",  "start": 6,  "end": 15, "hours": 9,  "code": "P"},
    "S2": {"label": "Siang", "start": 15, "end": 24, "hours": 9,  "code": "S"},
    "M":  {"label": "Malam", "start": 22, "end": 30, "hours": 8,  "code": "M"},
}

TEAMS    = ["A", "B", "C", "D"]
SECTORS  = [1, 2]
FACILITY = "t"

LEAVE_TYPES = {
    "CUTI_TAHUNAN":        "CUTI",
    "CUTI_ALASAN_PENTING": "CUTI",
    "SAKIT":               "SAKIT",
    "DIKLAT":              "DIKLAT",
    "MEDEC":               "MEDEC",
    "IELP":                "IELP",
}

JENIS_PENGAJUAN = {
    "Cuti Tahunan":        "CUTI_TAHUNAN",
    "Cuti Alasan Penting": "CUTI_ALASAN_PENTING",
    "Sakit":               "SAKIT",
    "Diklat":              "DIKLAT",
    "MEDEC":               "MEDEC",
    "IELP":                "IELP",
}

DISPLAY_CODE = {
    "S1": "PA", "S2": "SA", "M": "MA",
    "OFF": "L", "CUTI": "CT", "SAKIT": "SK",
    "DIKLAT": "DK", "MEDEC": "MC", "IELP": "IL",
}

ASSIGNABLE = ["S1", "S2", "OFF"]

RULES = {
    "max_consecutive_days":           6,
    "min_rest_after_block_hours":     60,
    "min_rest_between_shifts_hours":  11,
    "max_shift_hours":                9,
    "max_weekly_hours":               60,
    "max_time_in_position_hours":     3,
    "rest_break_minutes":             45,
    "min_staff_per_shift":            2,

    "max_night_shift_hours":          10,
    "min_rest_after_1_night_hours":   30,
    "min_rest_after_2_nights_hours":  54,
    "night_cannot_follow_s2_same_day": True,

    "max_shift_imbalance":            2,
    "imbalance_weight":               3,
}

HS_PARAMS = {
    "hm_size":      10,
    "hmcr":         0.80,
    "par":          0.30,
    "max_iter":     1000,
    "penalty_hard": 1000,
    "penalty_soft": 1,
}

DAY_NAMES_ID       = ["SEN", "SEL", "RAB", "KAM", "JUM", "SAB", "MIN"]
DAY_NAMES_ID_SHORT = ["Sn", "Sl", "Rb", "Km", "Jm", "Sb", "Mg"]
MONTH_NAMES_ID     = ["", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
                      "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]

BASE_DIR   = Path(__file__).resolve().parent
DATA_DIR   = BASE_DIR / "Data"
FILES_DIR  = DATA_DIR / "files"

PATH_GRID      = DATA_DIR / "atc_roster_grid.csv"
PATH_LEAVE     = DATA_DIR / "atc_leave_plan.csv"
PATH_PENGAJUAN = DATA_DIR / "pengurangan_hk.json"
PATH_STATE     = DATA_DIR / "state.json"
PATH_OUT_CSV   = FILES_DIR / "atc_roster_output.csv"
PATH_OUT_XLSX  = FILES_DIR / "atc_roster_output.xlsx"

DATA_DIR.mkdir(exist_ok=True)
FILES_DIR.mkdir(exist_ok=True)


def set_period(month: int, year: int) -> None:
    global MONTH, YEAR, DAYS
    MONTH = month
    YEAR  = year
    DAYS  = calendar.monthrange(year, month)[1]



def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(col).strip().upper() for col in df.columns]
    return df


def load_grid(path: Path) -> pd.DataFrame:
    df_grid = normalize_columns(pd.read_csv(path).fillna(""))
    for col in ("EMP_ID", "NAMA", "INITIAL", "SEKTOR"):
        if col not in df_grid.columns:
            raise ValueError(f"Kolom {col} tidak ada di file roster grid")
    return df_grid


def load_leave(path: Path) -> pd.DataFrame:
    df_leave = normalize_columns(pd.read_csv(path))
    for col in ("EMP_ID", "HARI_KE", "JENIS"):
        if col not in df_leave.columns:
            raise ValueError(f"Kolom {col} tidak ada di file leave plan")
    return df_leave


def build_employee_list(df_grid: pd.DataFrame) -> list[dict]:
    employees = []
    for _, row in df_grid.iterrows():
        employees.append({
            "id":      str(row["EMP_ID"]),
            "name":    row["NAMA"],
            "initial": row["INITIAL"],
            "sector":  row["SEKTOR"],
        })
    return employees


def build_locked_days(df_leave: pd.DataFrame) -> dict[str, dict[int, str]]:
    locked: dict[str, dict[int, str]] = {}
    for _, row in df_leave.iterrows():
        eid   = str(row["EMP_ID"])
        day   = int(row["HARI_KE"])
        jenis = LEAVE_TYPES.get(row["JENIS"], row["JENIS"])
        if 1 <= day <= DAYS:
            locked.setdefault(eid, {})[day] = jenis
    return locked



def encode_airnav(shift: str, sector: int, team: str) -> str:
    if shift not in SHIFT_DEF:
        return shift

    shift_code = SHIFT_DEF[shift]["code"]
    pos_type   = FACILITY
    return f"{shift_code}{pos_type}{sector}/{team}"


def decode_airnav(cell: str) -> tuple[str, int, str]:
    if "/" not in cell:
        return (cell, 0, "")

    prefix, team = cell.split("/")

    shift_code = prefix[0]
    sector     = int(prefix[2])

    shift_map  = {"P": "S1", "S": "S2", "M": "M"}
    shift      = shift_map.get(shift_code, shift_code)

    return (shift, sector, team)


def assign_team_rotation(shift: str, sector: int, day: int, emp_index: int) -> str:
    team_index = (day + emp_index + sector - 1) % len(TEAMS)
    team       = TEAMS[team_index]
    return encode_airnav(shift, sector, team)



def get_streak(schedule: dict, emp_id: str, up_to_day: int) -> int:
    streak = 0
    for d in range(up_to_day - 1, 0, -1):
        val = schedule[emp_id].get(d)
        if val in ("S1", "S2", "M"):
            streak += 1
        else:
            break
    return streak


def get_last_work_day(
    schedule: dict, emp_id: str, before_day: int
) -> tuple[Optional[int], Optional[str]]:
    for d in range(before_day - 1, 0, -1):
        val = schedule[emp_id].get(d)
        if val in ("S1", "S2", "M"):
            return d, val
    return None, None


def check_rest_gap(
    prev_shift: str, prev_day: int,
    new_shift:  str, new_day:  int
) -> bool:
    if new_shift == "M" and prev_shift == "S2" and prev_day == new_day:
        return False

    prev_end_abs  = prev_day * 24 + SHIFT_DEF[prev_shift]["end"]
    new_start_abs = new_day  * 24 + SHIFT_DEF[new_shift]["start"]
    gap           = new_start_abs - prev_end_abs

    if prev_shift == "M":
        return gap >= RULES["min_rest_after_1_night_hours"]

    return gap >= RULES["min_rest_between_shifts_hours"]


def check_weekly_hours(schedule: dict, emp_id: str, day: int) -> int:
    start = max(1, day - 6)
    total = 0
    for d in range(start, day + 1):
        val = schedule[emp_id].get(d)
        if val in SHIFT_DEF:
            total += SHIFT_DEF[val]["hours"]
    return total


def count_violations(schedule: dict, employees: list[dict]) -> dict:
    hard    = 0
    soft    = 0
    details = []

    for emp in employees:
        eid  = emp["id"]
        name = emp["name"]

        consecutive_nights = 0

        for day in range(1, DAYS + 1):
            val = schedule[eid].get(day)

            if val not in ("S1", "S2", "M"):
                consecutive_nights = 0
                continue

            streak = get_streak(schedule, eid, day)
            if streak + 1 > RULES["max_consecutive_days"]:
                hard += 1
                details.append(
                    f"[KERAS-H1] {name} hari-{day}: "
                    f"{streak+1} hari kerja berturut "
                    f"(maks {RULES['max_consecutive_days']})"
                )

            prev_day, prev_shift = get_last_work_day(schedule, eid, day)
            if prev_shift is not None:

                if not check_rest_gap(prev_shift, prev_day, val, day):
                    min_needed = (
                        RULES["min_rest_after_1_night_hours"]
                        if prev_shift == "M"
                        else RULES["min_rest_between_shifts_hours"]
                    )
                    actual_gap = (
                        day * 24 + SHIFT_DEF[val]["start"]
                    ) - (
                        prev_day * 24 + SHIFT_DEF[prev_shift]["end"]
                    )
                    hard += 1
                    details.append(
                        f"[KERAS-H2] {name} hari-{day}: "
                        f"jeda {actual_gap}j setelah {prev_shift} di hari-{prev_day} "
                        f"(min {min_needed}j)"
                    )

                if prev_shift == "M" and val == "M":
                    consecutive_nights += 1
                else:
                    consecutive_nights = 0

            if prev_shift == "M" and consecutive_nights >= 2:
                pp_day, pp_shift = get_last_work_day(schedule, eid, prev_day)
                if pp_shift == "M":
                    end_abs   = prev_day * 24 + SHIFT_DEF["M"]["end"]
                    start_abs = day * 24 + SHIFT_DEF[val]["start"]
                    gap       = start_abs - end_abs
                    if gap < RULES["min_rest_after_2_nights_hours"]:
                        hard += 1
                        details.append(
                            f"[KERAS-H3] {name} hari-{day}: "
                            f"jeda {gap}j setelah 2× malam berturut "
                            f"(min {RULES['min_rest_after_2_nights_hours']}j)"
                        )

            if day % 7 == 0 or day == DAYS:
                weekly = check_weekly_hours(schedule, eid, day)
                if weekly > RULES["max_weekly_hours"]:
                    soft += 1
                    details.append(
                        f"[LUNAK-S1] {name} minggu berakhir hari-{day}: "
                        f"{weekly}j (maks {RULES['max_weekly_hours']}j)"
                    )

    for day in range(1, DAYS + 1):
        counter = {"S1": 0, "S2": 0, "M": 0}
        for emp in employees:
            val = schedule[emp["id"]].get(day)
            if val in counter:
                counter[val] += 1

        for shift_key in ("S1", "S2"):
            count = counter[shift_key]
            if count < RULES["min_staff_per_shift"]:
                soft += (RULES["min_staff_per_shift"] - count)
                details.append(
                    f"[LUNAK-S2] hari-{day} shift {shift_key}: "
                    f"hanya {count} personel "
                    f"(min {RULES['min_staff_per_shift']})"
                )

        imbalance = abs(counter["S1"] - counter["S2"])
        if imbalance > RULES["max_shift_imbalance"]:
            excess = imbalance - RULES["max_shift_imbalance"]
            soft += excess * RULES["imbalance_weight"]
            details.append(
                f"[LUNAK-S3] hari-{day}: distribusi timpang "
                f"Pagi={counter['S1']} vs Siang={counter['S2']} "
                f"(selisih {imbalance}, maks {RULES['max_shift_imbalance']})"
            )

    total = hard * HS_PARAMS["penalty_hard"] + soft * HS_PARAMS["penalty_soft"]
    return {"hard": hard, "soft": soft, "total": total, "details": details}



def get_valid_shifts(
    schedule:    dict,
    emp_id:      str,
    day:         int,
    locked_days: dict
) -> list[str]:
    locked = locked_days.get(emp_id, {}).get(day)
    if locked:
        return [locked]

    streak = get_streak(schedule, emp_id, day)
    if streak >= RULES["max_consecutive_days"]:
        return ["OFF"]

    candidates = list(ASSIGNABLE)
    prev_day, prev_shift = get_last_work_day(schedule, emp_id, day)

    if prev_shift is not None:
        filtered = []
        for s in candidates:
            if s == "OFF":
                filtered.append(s)
            elif check_rest_gap(prev_shift, prev_day, s, day):
                filtered.append(s)
        candidates = filtered

    if not candidates:
        return ["OFF"]

    return candidates


def pick_shift(valid_shifts: list[str]) -> str:
    work_shifts = [s for s in valid_shifts if s in ("S1", "S2", "M")]

    if work_shifts and random.random() < 0.65:
        if "S1" in work_shifts and "S2" in work_shifts:
            return "S1" if random.random() < 0.65 else "S2"
        return random.choice(work_shifts)

    return random.choice(valid_shifts)


def pick_shift_balanced(valid_shifts: list[str], day_counts: dict) -> str:
    work_shifts = [s for s in valid_shifts if s in ("S1", "S2", "M")]

    if work_shifts and random.random() < 0.70:
        if "S1" in work_shifts and "S2" in work_shifts:
            s1, s2 = day_counts.get("S1", 0), day_counts.get("S2", 0)
            if s1 < s2:
                return "S1" if random.random() < 0.85 else "S2"
            if s2 < s1:
                return "S2" if random.random() < 0.85 else "S1"
            return random.choice(["S1", "S2"])
        return random.choice(work_shifts)

    return random.choice(valid_shifts)


def generate_harmony(employees: list[dict],
                     locked_days: dict) -> dict:
    schedule = {emp["id"]: {} for emp in employees}
    order = list(employees)
    for day in range(1, DAYS + 1):
        random.shuffle(order)
        day_counts = {"S1": 0, "S2": 0}
        for emp in order:
            valid = get_valid_shifts(schedule, emp["id"], day, locked_days)
            val = pick_shift_balanced(valid, day_counts)
            schedule[emp["id"]][day] = val
            if val in day_counts:
                day_counts[val] += 1
    return schedule


def initialize_harmony_memory(employees: list[dict],
                               locked_days: dict) -> list[dict]:
    hm = []
    for _ in range(HS_PARAMS["hm_size"]):
        sched = generate_harmony(employees, locked_days)
        score = count_violations(sched, employees)
        hm.append({"schedule": sched, "score": score})
    return sorted(hm, key=lambda x: x["score"]["total"])


def improvise_new_harmony(harmony_memory: list[dict],
                          employees: list[dict],
                          locked_days: dict) -> dict:
    new_sched = {emp["id"]: {} for emp in employees}

    for emp in employees:
        eid = emp["id"]
        for day in range(1, DAYS + 1):

            locked = locked_days.get(eid, {}).get(day)
            if locked:
                new_sched[eid][day] = locked
                continue

            if random.random() < HS_PARAMS["hmcr"]:
                src = random.choice(harmony_memory)
                val = src["schedule"][eid][day]

                if random.random() < HS_PARAMS["par"]:
                    val = "S2" if val == "S1" else ("S1" if val == "S2" else val)
            else:
                valid = get_valid_shifts(new_sched, eid, day, locked_days)
                val = pick_shift(valid)

            valid = get_valid_shifts(new_sched, eid, day, locked_days)
            if val not in valid:
                val = pick_shift(valid)

            new_sched[eid][day] = val

    return new_sched


def update_harmony_memory(harmony_memory: list[dict],
                          new_harmony: dict,
                          employees: list[dict]) -> tuple[list[dict], bool]:
    new_score = count_violations(new_harmony, employees)
    worst     = harmony_memory[-1]

    if new_score["total"] < worst["score"]["total"]:
        harmony_memory[-1] = {"schedule": new_harmony, "score": new_score}
        harmony_memory.sort(key=lambda x: x["score"]["total"])
        return harmony_memory, True

    return harmony_memory, False


def run_harmony_search(employees: list[dict],
                       locked_days: dict,
                       max_iter: Optional[int] = None) -> dict:
    hm = initialize_harmony_memory(employees, locked_days)
    best_score = hm[0]["score"]["total"]

    for i in range(max_iter or HS_PARAMS["max_iter"]):
        new_h = improvise_new_harmony(hm, employees, locked_days)
        hm, replaced = update_harmony_memory(hm, new_h, employees)

        if replaced and hm[0]["score"]["total"] < best_score:
            best_score = hm[0]["score"]["total"]

        if best_score == 0:
            break

    return hm[0]


def schedule_to_airnav(schedule: dict,
                       employees: list[dict]) -> dict:
    airnav = {}
    for i, emp in enumerate(employees):
        eid = emp["id"]
        airnav[eid] = {}
        for day in range(1, DAYS + 1):
            val = schedule[eid][day]
            if val in SHIFT_DEF:
                sector = SECTORS[(day + i) % len(SECTORS)]
                airnav[eid][day] = assign_team_rotation(val, sector, day, i)
            else:
                airnav[eid][day] = val
    return airnav



def day_name_map() -> dict[int, str]:
    return {
        d: DAY_NAMES_ID[date(YEAR, MONTH, d).weekday()]
        for d in range(1, DAYS + 1)
    }


def build_output_dataframe(airnav_schedule: dict,
                           employees: list[dict],
                           day_names: dict) -> pd.DataFrame:
    rows = []
    for emp in employees:
        eid = emp["id"]
        row = {
            "EMP_ID":  eid,
            "NAMA":    emp["name"],
            "INITIAL": emp["initial"],
            "SEKTOR":  emp["sector"],
        }

        total_hours = s1_cnt = s2_cnt = off_cnt = leave_cnt = 0

        for day in range(1, DAYS + 1):
            cell = airnav_schedule[eid][day]
            col  = f"{day} {day_names[day]}"
            row[col] = cell

            shift, _, _ = decode_airnav(cell)
            if shift in SHIFT_DEF:
                total_hours += SHIFT_DEF[shift]["hours"]
                if shift == "S1":   s1_cnt += 1
                elif shift == "S2": s2_cnt += 1
            elif shift == "OFF":
                off_cnt += 1
            else:
                leave_cnt += 1

        row["TOTAL_JAM"]  = total_hours
        row["JML_S1"]     = s1_cnt
        row["JML_S2"]     = s2_cnt
        row["JML_OFF"]    = off_cnt
        row["JML_CUTI"]   = leave_cnt
        rows.append(row)

    return pd.DataFrame(rows)


def export_to_excel(df_output: pd.DataFrame,
                    result: dict,
                    employees: list[dict],
                    output_path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_cell   = Font(name="Arial", size=9)
    font_bold   = Font(name="Arial", size=9, bold=True)

    fill_header = PatternFill("solid", start_color="1F4E79")
    fill_pagi   = PatternFill("solid", start_color="D6E4F0")
    fill_siang  = PatternFill("solid", start_color="D5E8D4")
    fill_malam  = PatternFill("solid", start_color="E1D5E7")
    fill_off    = PatternFill("solid", start_color="F2F2F2")
    fill_cuti   = PatternFill("solid", start_color="FFF2CC")
    fill_diklat = PatternFill("solid", start_color="FCE4D6")
    fill_sakit  = PatternFill("solid", start_color="F8CBAD")

    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_fill(value: str):
        if not isinstance(value, str):
            return None
        if value.startswith("P"):  return fill_pagi
        if value.startswith("S") and "/" in value: return fill_siang
        if value.startswith("M") and "/" in value: return fill_malam
        if value == "OFF":         return fill_off
        if value == "CUTI":        return fill_cuti
        if value == "DIKLAT":      return fill_diklat
        if value in ("SAKIT", "MEDEC", "IELP"): return fill_sakit
        return None

    ws = wb.active
    ws.title = "Roster"

    ncols = len(df_output.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ncols, 20))
    title_cell = ws.cell(row=1, column=1,
                         value=f"JADWAL DINAS ATC TWR — {CATEGORY} KATEGORI — {MONTH:02d}/{YEAR}")
    title_cell.font = Font(name="Arial", size=12, bold=True)
    title_cell.alignment = center

    header_row = 3
    for c_idx, col_name in enumerate(df_output.columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = border

    for r_idx, (_, row) in enumerate(df_output.iterrows(), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = font_cell
            cell.alignment = center
            cell.border = border
            f = cell_fill(value)
            if f:
                cell.fill = f

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    for c in range(5, 5 + DAYS):
        ws.column_dimensions[get_column_letter(c)].width = 7.5
    for c in range(5 + DAYS, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    ws.freeze_panes = "E4"

    ws2 = wb.create_sheet("Pelanggaran")
    ws2.cell(row=1, column=1, value="LAPORAN PELANGGARAN").font = Font(
        name="Arial", size=12, bold=True)
    ws2.cell(row=3, column=1, value="Jenis").font = font_bold
    ws2.cell(row=3, column=2, value="Jumlah").font = font_bold
    ws2.cell(row=4, column=1, value="KERAS")
    ws2.cell(row=4, column=2, value=result["score"]["hard"])
    ws2.cell(row=5, column=1, value="LUNAK")
    ws2.cell(row=5, column=2, value=result["score"]["soft"])
    ws2.cell(row=6, column=1, value="Skor Total")
    ws2.cell(row=6, column=2, value=result["score"]["total"])

    ws2.cell(row=8, column=1, value="Detail:").font = font_bold
    for i, detail in enumerate(result["score"]["details"], start=9):
        ws2.cell(row=i, column=1, value=detail).font = font_cell
    ws2.column_dimensions["A"].width = 80
    ws2.column_dimensions["B"].width = 12

    ws3 = wb.create_sheet("Ringkasan")
    headers = ["EMP_ID", "NAMA", "TOTAL_JAM", "JML_S1", "JML_S2",
               "JML_OFF", "JML_CUTI", "PERSENTASE_KERJA"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center

    n_emp = len(df_output)
    col_total_jam = get_column_letter(5 + DAYS)
    col_s1        = get_column_letter(5 + DAYS + 1)
    col_s2        = get_column_letter(5 + DAYS + 2)
    col_off       = get_column_letter(5 + DAYS + 3)
    col_cuti      = get_column_letter(5 + DAYS + 4)

    for i in range(n_emp):
        roster_row = header_row + 1 + i
        r = i + 2
        ws3.cell(row=r, column=1, value=f"=Roster!A{roster_row}")
        ws3.cell(row=r, column=2, value=f"=Roster!B{roster_row}")
        ws3.cell(row=r, column=3, value=f"=Roster!{col_total_jam}{roster_row}")
        ws3.cell(row=r, column=4, value=f"=Roster!{col_s1}{roster_row}")
        ws3.cell(row=r, column=5, value=f"=Roster!{col_s2}{roster_row}")
        ws3.cell(row=r, column=6, value=f"=Roster!{col_off}{roster_row}")
        ws3.cell(row=r, column=7, value=f"=Roster!{col_cuti}{roster_row}")
        ws3.cell(row=r, column=8, value=f"=(D{r}+E{r})/{DAYS}")
        ws3.cell(row=r, column=8).number_format = "0.0%"

    total_r = n_emp + 2
    ws3.cell(row=total_r, column=2, value="TOTAL").font = font_bold
    for col, letter in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")]:
        ws3.cell(row=total_r, column=col,
                 value=f"=SUM({letter}2:{letter}{n_emp+1})").font = font_bold

    for c, w in [("A", 8), ("B", 18), ("C", 11), ("D", 8), ("E", 8),
                 ("F", 9), ("G", 10), ("H", 16)]:
        ws3.column_dimensions[c].width = w

    wb.save(output_path)



def load_pengajuan() -> list[dict]:
    if PATH_PENGAJUAN.exists():
        return json.loads(PATH_PENGAJUAN.read_text())
    return []


def save_pengajuan(items: list[dict]) -> None:
    PATH_PENGAJUAN.write_text(json.dumps(items, indent=2, ensure_ascii=False))


def pengajuan_to_locked_days(items: list[dict]) -> dict[str, dict[int, str]]:
    locked: dict[str, dict[int, str]] = {}
    for item in items:
        eid   = str(item["emp_id"])
        jenis = LEAVE_TYPES.get(JENIS_PENGAJUAN.get(item["jenis"], item["jenis"]),
                                "CUTI")
        start = date.fromisoformat(item["tanggal_mulai"])
        end   = date.fromisoformat(item["tanggal_selesai"])
        d = start
        while d <= end:
            if d.year == YEAR and d.month == MONTH:
                locked.setdefault(eid, {})[d.day] = jenis
            d += timedelta(days=1)
    return locked


def merge_locked(base: dict, extra: dict) -> dict:
    merged = {eid: dict(days) for eid, days in base.items()}
    for eid, days in extra.items():
        merged.setdefault(eid, {}).update(days)
    return merged


def save_state(employees: list[dict], result: dict, airnav: dict) -> None:
    state = {
        "month":        MONTH,
        "year":         YEAR,
        "days":         DAYS,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "employees":    employees,
        "score":        result["score"],
        "schedule":     {eid: {str(d): v for d, v in days.items()}
                         for eid, days in result["schedule"].items()},
        "airnav":       {eid: {str(d): v for d, v in days.items()}
                         for eid, days in airnav.items()},
    }
    PATH_STATE.write_text(json.dumps(state, indent=1, ensure_ascii=False))


def load_state() -> Optional[dict]:
    if not PATH_STATE.exists():
        return None
    state = json.loads(PATH_STATE.read_text())
    state["schedule"] = {eid: {int(d): v for d, v in days.items()}
                         for eid, days in state["schedule"].items()}
    state["airnav"]   = {eid: {int(d): v for d, v in days.items()}
                         for eid, days in state["airnav"].items()}
    return state


def regenerate(max_iter: Optional[int] = None,
               seed: Optional[int] = None) -> dict:
    if seed is not None:
        random.seed(seed)
        np.random.seed(seed % (2**32))

    if not PATH_GRID.exists():
        raise FileNotFoundError(
            "Belum ada data personel. Upload file roster grid lewat halaman Input Jadwal.")

    df_grid   = load_grid(PATH_GRID)
    employees = build_employee_list(df_grid)

    locked = {}
    if PATH_LEAVE.exists():
        locked = build_locked_days(load_leave(PATH_LEAVE))
    locked = merge_locked(locked, pengajuan_to_locked_days(load_pengajuan()))

    result = run_harmony_search(employees, locked, max_iter=max_iter)
    airnav = schedule_to_airnav(result["schedule"], employees)

    df_output = build_output_dataframe(airnav, employees, day_name_map())
    df_output.to_csv(PATH_OUT_CSV, index=False)
    export_to_excel(df_output, result, employees, PATH_OUT_XLSX)

    save_state(employees, result, airnav)
    return load_state()


def jenis_to_internal_value(jenis_display: str) -> str:
    """Nama jenis di form ('Cuti Tahunan') → nilai internal jadwal ('CUTI')."""
    kode = JENIS_PENGAJUAN.get(jenis_display, jenis_display)
    return LEAVE_TYPES.get(kode, "CUTI")


def set_state_days(state: dict, emp_id: str,
                   start: date, end: date, value: str) -> None:
    """Timpa sel jadwal seorang personel utk rentang tanggal dgn `value`
    (mis. 'CUTI'/'SAKIT' utk libur, atau 'OFF' saat pengajuan dibatalkan).
    Hanya menyentuh personel ybs — personel lain tidak berubah."""
    if emp_id not in state["schedule"]:
        return
    d = start
    while d <= end:
        if (d.year, d.month) == (state["year"], state["month"]):
            day = d.day
            state["schedule"][emp_id][day] = value
            state["airnav"][emp_id][day]   = value
        d += timedelta(days=1)


def save_patched_state(state: dict) -> dict:
    """Hitung ulang skor, tulis ulang file output, dan simpan state yang sudah
    di-patch in-place (tanpa menjalankan Harmony Search lagi)."""
    set_period(state["month"], state["year"])
    employees = state["employees"]
    schedule  = state["schedule"]
    airnav    = state["airnav"]

    score = count_violations(schedule, employees)
    result = {"schedule": schedule, "score": score}

    df_output = build_output_dataframe(airnav, employees, day_name_map())
    df_output.to_csv(PATH_OUT_CSV, index=False)
    export_to_excel(df_output, result, employees, PATH_OUT_XLSX)

    save_state(employees, result, airnav)
    return load_state()



app = Flask(__name__)
CORS(app)

app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

@app.route("/")
def serve_dashboard():
    return send_file(BASE_DIR / "resources" / "views" / "dashboard.blade.php", mimetype="text/html")

@app.route("/jadwal")
def serve_jadwal():
    return send_file(BASE_DIR / "resources" / "views" / "jadwal.blade.php", mimetype="text/html")

@app.route("/input-jadwal")
def serve_input_jadwal():
    return send_file(BASE_DIR / "resources" / "views" / "input-jadwal.blade.php", mimetype="text/html")

@app.route("/pengurangan-hk")
def serve_pengurangan_hk():
    return send_file(BASE_DIR / "resources" / "views" / "pengurangan-hk.blade.php", mimetype="text/html")

def require_state() -> dict:
    state = load_state()
    if state is None:
        raise FileNotFoundError(
            "Jadwal belum di-generate. Upload data lewat halaman Input Jadwal dulu.")
    set_period(state["month"], state["year"])
    return state


@app.errorhandler(FileNotFoundError)
@app.errorhandler(ValueError)
def handle_user_error(err):
    return jsonify({"error": str(err)}), 400


@app.get("/api/status")
def api_status():
    state = load_state()
    return jsonify({
        "has_data":     PATH_GRID.exists(),
        "has_schedule": state is not None,
        "month":        state["month"] if state else MONTH,
        "year":         state["year"] if state else YEAR,
        "days":         state["days"] if state else DAYS,
        "month_name":   MONTH_NAMES_ID[state["month"] if state else MONTH],
        "category":     CATEGORY,
        "generated_at": state["generated_at"] if state else None,
        "employee_count": len(state["employees"]) if state else 0,
        "score":        state["score"] if state else None,
        "pengajuan_count": len(load_pengajuan()),
    })


@app.get("/api/employees")
def api_employees():
    state = require_state()
    return jsonify({"employees": state["employees"]})



@app.post("/api/upload")
def api_upload():
    grid_file  = request.files.get("grid") or request.files.get("file")
    leave_file = request.files.get("leave")

    if grid_file is None and leave_file is None:
        raise ValueError("Tidak ada file yang diupload. Sertakan file 'grid' "
                         "(personel) dan/atau 'leave' (rencana cuti).")

    month = int(request.form.get("month", MONTH))
    year  = int(request.form.get("year", YEAR))
    if not 1 <= month <= 12:
        raise ValueError("Bulan tidak valid (1-12).")
    set_period(month, year)

    if grid_file is not None:
        content = grid_file.read()
        df = load_grid(io.BytesIO(content))
        if df.empty:
            raise ValueError("File roster grid kosong.")
        PATH_GRID.write_bytes(content)

    if leave_file is not None:
        content = leave_file.read()
        load_leave(io.BytesIO(content))
        PATH_LEAVE.write_bytes(content)

    state = regenerate()
    return jsonify({
        "message":      "File diterima, jadwal berhasil di-generate.",
        "month":        state["month"],
        "year":         state["year"],
        "days":         state["days"],
        "employee_count": len(state["employees"]),
        "score":        state["score"],
    })


@app.post("/api/generate")
def api_generate():
    max_iter = request.args.get("iterations", type=int)
    prev = require_state()
    prev_airnav = prev["airnav"]

    state = prev
    for _ in range(4):
        seed = uuid.uuid4().int % (2**31)
        state = regenerate(max_iter=max_iter, seed=seed)
        if state["airnav"] != prev_airnav:
            break

    return jsonify({
        "message": "Jadwal baru berhasil di-generate.",
        "score":   state["score"],
        "generated_at": state["generated_at"],
    })



@app.get("/api/schedule/day")
def api_schedule_day():
    state = require_state()
    day = request.args.get("day", default=date.today().day, type=int)
    if not 1 <= day <= state["days"]:
        raise ValueError(f"Hari harus 1-{state['days']}.")

    shift_filter = request.args.get("shift")

    emp_by_id = {e["id"]: e for e in state["employees"]}

    grid: dict[int, dict[str, dict]] = {
        s: {k: {"codes": [], "personnel": []} for k in SHIFT_DEF} for s in SECTORS
    }
    off, leave = [], []

    for eid, days in state["airnav"].items():
        cell = days[day]
        emp  = emp_by_id[eid]
        person = {"emp_id": eid, "nama": emp["name"], "initial": emp["initial"]}

        shift, sector, team = decode_airnav(cell)
        if shift in SHIFT_DEF:
            if shift_filter and shift != shift_filter:
                continue
            slot = grid[sector][shift]
            if cell not in slot["codes"]:
                slot["codes"].append(cell)
            slot["personnel"].append({**person, "code": cell, "team": team})
        elif cell == "OFF":
            off.append(person)
        else:
            leave.append({**person, "jenis": cell})

    rows = []
    for sector in SECTORS:
        row = {"sector": f"{FACILITY}{sector}"}
        for key, sdef in SHIFT_DEF.items():
            slot = grid[sector][key]
            row[sdef["label"].lower()] = {
                "codes":     slot["codes"],
                "personnel": slot["personnel"],
                "count":     len(slot["personnel"]),
            }
        rows.append(row)

    d = date(state["year"], state["month"], day)
    return jsonify({
        "day":      day,
        "date":     d.isoformat(),
        "day_name": DAY_NAMES_ID[d.weekday()],
        "month_name": MONTH_NAMES_ID[state["month"]],
        "shift_def": {k: {"label": v["label"],
                          "time": f"{v['start']:02d}:00 - {v['end'] % 24:02d}:00"}
                      for k, v in SHIFT_DEF.items()},
        "grid":  rows,
        "off":   off,
        "leave": leave,
    })



@app.get("/api/schedule/month")
def api_schedule_month():
    state = require_state()

    headers = []
    for d in range(1, state["days"] + 1):
        wd = date(state["year"], state["month"], d).weekday()
        headers.append({
            "day":     d,
            "name":    DAY_NAMES_ID_SHORT[wd],
            "weekend": wd >= 5,
        })

    rows = []
    for no, emp in enumerate(state["employees"], start=1):
        eid   = emp["id"]
        cells = []
        hk    = 0
        for d in range(1, state["days"] + 1):
            airnav_cell = state["airnav"][eid][d]
            raw         = state["schedule"][eid][d]
            code        = DISPLAY_CODE.get(raw, raw)
            if raw in SHIFT_DEF:
                hk += 1
            cells.append({"day": d, "code": code, "airnav": airnav_cell})
        rows.append({
            "no":      no,
            "emp_id":  eid,
            "nama":    emp["name"],
            "initial": emp["initial"],
            "cells":   cells,
            "hk":      hk,
        })

    return jsonify({
        "title":   f"JADWAL DINAS ATC — {MONTH_NAMES_ID[state['month']]} "
                   f"{state['year']} ({CATEGORY} KATEGORI)",
        "month":   state["month"],
        "year":    state["year"],
        "days":    state["days"],
        "shift_info": f"Shift I: 06:00-15:00 LT (9 jam)  |  "
                      f"Shift II: 15:00-24:00 LT (9 jam)  |  "
                      f"Personel per shift: {RULES['min_staff_per_shift']}",
        "day_headers": headers,
        "rows":    rows,
        "legend":  DISPLAY_CODE,
        "score":   state["score"],
    })



@app.get("/api/export")
def api_export():
    state = require_state()
    fmt = request.args.get("format", "xlsx").lower()
    day = request.args.get("day", type=int)

    if day is not None:
        if not 1 <= day <= state["days"]:
            raise ValueError(f"Hari harus 1-{state['days']}.")
        rows = []
        for emp in state["employees"]:
            rows.append({
                "EMP_ID":  emp["id"],
                "NAMA":    emp["name"],
                "INITIAL": emp["initial"],
                "JADWAL":  state["airnav"][emp["id"]][day],
            })
        buf = io.BytesIO()
        pd.DataFrame(rows).to_csv(buf, index=False)
        buf.seek(0)
        fname = f"jadwal_atc_{state['year']}-{state['month']:02d}-{day:02d}.csv"
        return send_file(buf, mimetype="text/csv",
                         as_attachment=True, download_name=fname)

    if fmt == "csv":
        if not PATH_OUT_CSV.exists():
            raise FileNotFoundError("File output belum ada, generate jadwal dulu.")
        return send_file(PATH_OUT_CSV, mimetype="text/csv", as_attachment=True,
                         download_name=f"atc_roster_{state['year']}-{state['month']:02d}.csv")

    if not PATH_OUT_XLSX.exists():
        raise FileNotFoundError("File output belum ada, generate jadwal dulu.")
    return send_file(
        PATH_OUT_XLSX,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"atc_roster_{state['year']}-{state['month']:02d}.xlsx")



def format_pengajuan(item: dict, no: int, emp_by_id: dict) -> dict:
    start = date.fromisoformat(item["tanggal_mulai"])
    end   = date.fromisoformat(item["tanggal_selesai"])
    emp   = emp_by_id.get(str(item["emp_id"]), {})
    return {
        "no":              no,
        "id":              item.get("id"),
        "emp_id":          item["emp_id"],
        "nama":            emp.get("name", item.get("nama", "?")),
        "initial":         emp.get("initial", ""),
        "jenis":           item["jenis"],
        "tanggal_mulai":   item["tanggal_mulai"],
        "tanggal_selesai": item["tanggal_selesai"],
        "durasi":          f"{(end - start).days + 1} Hari",
        "source":          item.get("source", "manual"),
    }


def leave_plan_items() -> list[dict]:
    if not PATH_LEAVE.exists():
        return []

    df = load_leave(PATH_LEAVE)
    code_to_display = {v: k for k, v in JENIS_PENGAJUAN.items()}

    groups: dict[tuple, list[int]] = {}
    for _, row in df.iterrows():
        eid  = str(row["EMP_ID"])
        day  = int(row["HARI_KE"])
        kode = str(row["JENIS"])
        if 1 <= day <= DAYS:
            groups.setdefault((eid, kode), []).append(day)

    items = []
    for (eid, kode), days in groups.items():
        days.sort()
        run_start = prev = days[0]
        for d in days[1:] + [None]:
            if d == prev + 1:
                prev = d
            else:
                items.append({
                    "emp_id":          eid,
                    "jenis":           code_to_display.get(kode, kode),
                    "tanggal_mulai":   date(YEAR, MONTH, run_start).isoformat(),
                    "tanggal_selesai": date(YEAR, MONTH, prev).isoformat(),
                    "source":          "leave_plan",
                })
                if d is not None:
                    run_start = prev = d
    items.sort(key=lambda x: (x["tanggal_mulai"], x["emp_id"]))
    return items


def validate_pengajuan_payload(data: dict, employees: list[dict]) -> dict:
    for field in ("emp_id", "jenis", "tanggal_mulai", "tanggal_selesai"):
        if not data.get(field):
            raise ValueError(f"Field '{field}' wajib diisi.")

    emp_ids = {e["id"] for e in employees}
    if str(data["emp_id"]) not in emp_ids:
        raise ValueError(f"Personel {data['emp_id']} tidak ditemukan.")

    if data["jenis"] not in JENIS_PENGAJUAN:
        raise ValueError(f"Jenis tidak valid. Pilihan: {', '.join(JENIS_PENGAJUAN)}.")

    try:
        start = date.fromisoformat(data["tanggal_mulai"])
        end   = date.fromisoformat(data["tanggal_selesai"])
    except ValueError:
        raise ValueError("Format tanggal harus YYYY-MM-DD.")

    if end < start:
        raise ValueError("Tanggal selesai tidak boleh sebelum tanggal mulai.")
    if (start.year, start.month) != (YEAR, MONTH) and (end.year, end.month) != (YEAR, MONTH):
        raise ValueError(
            f"Rentang tanggal harus menyentuh bulan aktif "
            f"({MONTH_NAMES_ID[MONTH].title()} {YEAR}).")

    return {
        "emp_id":          str(data["emp_id"]),
        "jenis":           data["jenis"],
        "tanggal_mulai":   data["tanggal_mulai"],
        "tanggal_selesai": data["tanggal_selesai"],
    }


@app.get("/api/pengurangan-hk")
def api_pengajuan_list():
    state = load_state()
    if state:
        set_period(state["month"], state["year"])
    emp_by_id = {e["id"]: e for e in (state["employees"] if state else [])}

    items = leave_plan_items() + load_pengajuan()
    return jsonify({
        "jenis_options": list(JENIS_PENGAJUAN.keys()),
        "items": [format_pengajuan(it, i + 1, emp_by_id)
                  for i, it in enumerate(items)],
    })


@app.post("/api/pengurangan-hk")
def api_pengajuan_add():
    state = require_state()
    payload = validate_pengajuan_payload(request.get_json(force=True),
                                         state["employees"])
    payload["id"] = uuid.uuid4().hex[:8]

    items = load_pengajuan()
    items.append(payload)
    save_pengajuan(items)

    # Keluarkan personel dari shift pada tanggal cuti & tandai tidak tersedia,
    # tanpa men-generate ulang seluruh roster (personel lain tetap).
    set_state_days(state, payload["emp_id"],
                   date.fromisoformat(payload["tanggal_mulai"]),
                   date.fromisoformat(payload["tanggal_selesai"]),
                   jenis_to_internal_value(payload["jenis"]))
    state = save_patched_state(state)

    emp_by_id = {e["id"]: e for e in state["employees"]}
    return jsonify({
        "message": "Pengajuan ditambahkan. Personel dikeluarkan dari shift "
                   "pada tanggal tersebut.",
        "item":    format_pengajuan(payload, len(items), emp_by_id),
        "score":   state["score"],
    }), 201


@app.put("/api/pengurangan-hk/<item_id>")
def api_pengajuan_update(item_id: str):
    state = require_state()
    items = load_pengajuan()
    idx = next((i for i, it in enumerate(items) if it["id"] == item_id), None)
    if idx is None:
        raise ValueError(f"Pengajuan {item_id} tidak ditemukan.")

    old = items[idx]
    payload = validate_pengajuan_payload(request.get_json(force=True),
                                         state["employees"])
    payload["id"] = item_id
    items[idx] = payload
    save_pengajuan(items)

    # Kembalikan hari cuti lama jadi OFF, lalu terapkan rentang cuti yg baru.
    set_state_days(state, old["emp_id"],
                   date.fromisoformat(old["tanggal_mulai"]),
                   date.fromisoformat(old["tanggal_selesai"]),
                   "OFF")
    set_state_days(state, payload["emp_id"],
                   date.fromisoformat(payload["tanggal_mulai"]),
                   date.fromisoformat(payload["tanggal_selesai"]),
                   jenis_to_internal_value(payload["jenis"]))
    state = save_patched_state(state)

    emp_by_id = {e["id"]: e for e in state["employees"]}
    return jsonify({
        "message": "Pengajuan diperbarui. Jadwal personel terkait disesuaikan.",
        "item":    format_pengajuan(payload, idx + 1, emp_by_id),
        "score":   state["score"],
    })


@app.delete("/api/pengurangan-hk/<item_id>")
def api_pengajuan_delete(item_id: str):
    state = require_state()
    items = load_pengajuan()
    removed = next((it for it in items if it["id"] == item_id), None)
    if removed is None:
        raise ValueError(f"Pengajuan {item_id} tidak ditemukan.")
    save_pengajuan([it for it in items if it["id"] != item_id])

    # Bebaskan hari cuti tsb jadi OFF (personel tidak otomatis dapat shift lagi;
    # gunakan 'Generate Jadwal Baru' bila ingin mengisinya kembali).
    set_state_days(state, removed["emp_id"],
                   date.fromisoformat(removed["tanggal_mulai"]),
                   date.fromisoformat(removed["tanggal_selesai"]),
                   "OFF")
    state = save_patched_state(state)

    return jsonify({
        "message": "Pengajuan dihapus. Hari cuti dibebaskan menjadi OFF.",
        "score":   state["score"],
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
